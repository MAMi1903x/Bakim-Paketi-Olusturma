import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl import load_workbook

# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="Bakım Paketi Oluşturucu", layout="wide")
st.title("✈️ Bakım Paketi Excel Oluşturucu (Production)")

pdf_file = st.file_uploader("PDF Yükle", type=["pdf"])
template_file = st.file_uploader("Excel Şablon Yükle", type=["xlsx"])
wo_number = st.text_input("W/O Numarası")

use_engineering = st.checkbox("Mühendislik değerlendirmesi kullan (opsiyonel)")
map_file = None
if use_engineering:
    map_file = st.file_uploader(
        "Mühendislik Değerlendirmesi Excel Yükle (DESCRIPTION + CMT/IMT/CDCCL/KOMPLEKS)",
        type=["xlsx"]
    )

# -----------------------------
# Session state
# -----------------------------
if "filled_xlsx" not in st.session_state:
    st.session_state["filled_xlsx"] = None
if "filled_tsv" not in st.session_state:
    st.session_state["filled_tsv"] = None
if "dl_aircraft" not in st.session_state:
    st.session_state["dl_aircraft"] = ""
if "dl_v" not in st.session_state:
    st.session_state["dl_v"] = 0

# -----------------------------
# Helpers
# -----------------------------
def mpd_cmr_interval(tasks, target_card="52-360-00-01") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def NDT_control(tasks, target_card="EOD-B737-53-0010") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def access_issue(tasks, target_card="EOD-B737-51-0010") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def access_issue2(tasks, target_card="55-826-01-01") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def access_issue3(tasks, target_card="55-840-02-01") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def has_eod_max_engine_run_card(tasks, target_card="EOD-B737-73-0003") -> bool:
    target = str(target_card).strip().upper()
    for t in tasks:
        text_to_check = " ".join([
            str(t.get("match_key", "")),
            str(t.get("description", "")),
            str(t.get("row_text", "")),
            str(t.get("card_no", "")),
        ]).upper()

        if target in text_to_check:
            return True
    return False
def get_location_from_package(package_name: str) -> str:
    package_name = (package_name or "").strip().upper()
    return package_name[-3:] if len(package_name) >= 3 else ""

def normalize_skill(skill_value):
    s = str(skill_value).strip().upper()
    return s if s in ("B1", "B2") else "B1"

def norm_header(s) -> str:
    return str(s).strip().lower() if s is not None else ""

def clean_text_key(text) -> str:
    if text is None:
        return ""
    s = str(text).upper().replace("İ", "I")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def clean_description(text) -> str:
    return clean_text_key(text)

def yn_from_any(val) -> str:
    if val is None:
        return "N"
    s = str(val).strip().upper()
    return "Y" if s in ("Y", "YES", "TRUE", "1", "T") else "N"

def parse_mh_and_skill(value):
    if value is None:
        return "", ""
    v = str(value).strip()
    if not v:
        return "", ""

    skill = ""
    mh_part = v

    if "/" in v:
        mh_part, skill = v.split("/", 1)
        mh_part = mh_part.strip()
        skill = normalize_skill(skill)

    if ":" in mh_part:
        try:
            h = int(float(mh_part.split(":")[0]))
            return str(h), skill
        except Exception:
            return mh_part, skill

    try:
        return str(int(float(mh_part))), skill
    except Exception:
        return mh_part, skill

def safe_cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\t", " ").replace("\n", " ").replace("\r", " ")

def normalize_text_for_search(text: str) -> str:
    if not text:
        return ""
    txt = str(text).upper().replace("İ", "I")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()

def extract_card_no(text: str) -> str:
    if not text:
        return ""

    text = str(text).upper()

    m = re.search(r"\b\d{2}-\d{3}-\d{2}-\d{2}\b", text)
    if m:
        return m.group(0)

    m = re.search(r"\b\d{2}-\d{3}-\d{2}\b", text)
    if m:
        return m.group(0)

    return ""

def format_num(n):
    try:
        n = float(n)
        if n.is_integer():
            return str(int(n))
        return f"{n:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(n)

# -----------------------------
# Cover info
# -----------------------------
def extract_cover_info(full_text: str):
    package_name = ""
    aircraft = ""

    m_type = re.search(r"Type\s*Of\s*Work\s*:?\s*(.+)", full_text, re.IGNORECASE)
    if m_type:
        package_name = m_type.group(1).strip()

    m_reg = re.search(r"A/C Type\s*/\s*Registration\s*(.+)", full_text, re.IGNORECASE)
    if m_reg:
        aircraft = m_reg.group(1).split("/")[-1].strip()

    return aircraft, package_name

def detect_aircraft_family_from_cover(pdf_bytes: bytes):
    full_text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    ac_match = re.search(r"A/C Type\s*/\s*Registration\s*(.+)", full_text, re.IGNORECASE)
    if not ac_match:
        return None, "⚠️ A/C Type / Registration bulunamadı."

    ac_info = ac_match.group(1).strip()
    prefix4 = ac_info[:4].upper()

    if prefix4 == "B73N":
        return "B737NG", "✈️ Uçak tipi: B737NG (B73N)"
    if prefix4 == "B73M":
        return "B737MAX", "✈️ Uçak tipi: B737MAX (B73M) ‼️YETKİ KONTROLÜ YAPILMASI GEREKİYOR‼️"

    return "UNKNOWN", f"⚠️ Uçak tipi tanınamadı (ilk 4 hane: {prefix4})"

# -----------------------------
# Summary table detection
# -----------------------------
def is_summary_page(page_text: str) -> bool:
    return "SUMMARY" in (page_text or "").upper()

def normalize_colname(c) -> str:
    return str(c).strip().upper() if c is not None else ""

def find_best_columns(df_cols):
    desc_col = next((c for c in df_cols if "DESC" in normalize_colname(c)), None)
    mh_col = next((c for c in df_cols if "MH" in normalize_colname(c)), None)

    ref_col = next((c for c in df_cols if ("W/O" in normalize_colname(c) and "REFER" in normalize_colname(c))), None)
    if ref_col is None:
        ref_col = next((c for c in df_cols if "REFER" in normalize_colname(c)), None)
    if ref_col is None:
        ref_col = next((c for c in df_cols if ("W/O" in normalize_colname(c) or "WO" in normalize_colname(c))), None)

    return desc_col, mh_col, ref_col

def table_looks_like_summary(header_row) -> bool:
    header = [normalize_colname(h) for h in header_row]
    has_desc = any("DESC" in h for h in header)
    has_mh = any("MH" in h for h in header)
    has_ref = (
        any("REFER" in h for h in header)
        or any("W/O" in h for h in header)
        or any("WO" in h.replace(" ", "") for h in header)
    )
    return has_desc and has_mh and has_ref

def extract_summary_tasks(pdf_bytes: bytes):
    full_text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for p in pdf.pages:
            t = p.extract_text()
            if t:
                full_text += t + "\n"

    aircraft, package_name = extract_cover_info(full_text)
    camo_prefix = f"PLEASE PERFORM CAMO WP: {package_name} | "
    camo_prefix = camo_prefix.upper().replace("İ", "I")

    tasks = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""
            if not is_summary_page(page_text):
                continue

            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue

                header = table[0]
                if not table_looks_like_summary(header):
                    continue

                df = pd.DataFrame(table[1:], columns=header)
                desc_col, mh_col, ref_col = find_best_columns(df.columns)

                if desc_col is None or mh_col is None:
                    continue

                for _, row in df.iterrows():
                    row_text = " ".join([str(x) for x in row.values if x is not None])
                    row_text_clean = normalize_text_for_search(row_text)

                    raw_desc = clean_description(row.get(desc_col, ""))
                    if not raw_desc or raw_desc.lower() == "none":
                        continue

                    card_no = extract_card_no(row_text_clean)

                    mh, skill = parse_mh_and_skill(row.get(mh_col, ""))
                    if mh in ("", "0"):
                        mh = "1"

                    needs_wo = "-" not in raw_desc[:20]

                    wo_prefix = ""
                    if needs_wo and ref_col is not None:
                        ref_val = str(row.get(ref_col, "") or "")
                        m = re.search(r"\b\d{3,}\b", ref_val)
                        if m:
                            wo_prefix = f"WO:{m.group(0)} "

                    if needs_wo and wo_prefix:
                        final_desc = camo_prefix + wo_prefix
                    elif needs_wo and not wo_prefix:
                        final_desc = camo_prefix + raw_desc
                    else:
                        final_desc = camo_prefix + raw_desc

                    tasks.append({
                        "page_no": page_no,
                        "card_no": card_no,
                        "description": final_desc,
                        "match_key": raw_desc,
                        "row_text": row_text_clean,
                        "man_hour": mh,
                        "skill": skill,
                        "rII": "N",
                        "critical_task": "N",
                        "cdccl": "N",
                        "intervals": [],
                        "interval_exceed": "N",
                        "interval_summary": "",
                    })

    return aircraft, package_name, tasks

# -----------------------------
# Interval parsing
# -----------------------------
STOP_WORDS = {
    "ZONE", "ACCESS", "TASK", "WORK AREA", "SKILL", "APPLICABILITY",
    "REFERENCES", "NOTE", "MAN-HOURS", "MH EST", "DESCRIPTION",
    "METHOD", "END OF TASK", "CARD NUMBER", "BOEING CARD NO",
    "PLANNING INFORMATION", "A. EFFECTIVITY"
}

def interval_exceeds(interval_type, value, aircraft_family=None):
    t = str(interval_type).upper().strip()
    try:
        v = float(value)
    except Exception:
        return False

    family = (aircraft_family or "").upper().strip()

    # B737MAX kuralları
    if family == "B737MAX":
        if t == "FH":
            return v > 1600
        elif t == "DY":
            return v > 120
        elif t == "MO":
            return v > 4
        return False

    # Varsayılan / B737NG kuralları
    if t == "FH":
        return v >= 15000
    elif t == "FC":
        return v >= 4500
    elif t == "YR":
        return v >= 3
    return False

def convert_mo_to_yr(mo_value):
    try:
        return float(mo_value) / 12.0
    except Exception:
        return None

def extract_intervals_from_chunk(chunk_text, source_name, aircraft_family=None):
    results = []
    if not chunk_text:
        return results

    found = re.findall(r"(\d+(?:\.\d+)?)\s*(FH|FC|YR|MO|DY)\b", chunk_text, re.IGNORECASE)

    family = (aircraft_family or "").upper().strip()

    for value, typ in found:
        try:
            v = float(value)
        except Exception:
            continue

        t = typ.upper()

        # NG için MO -> YR dönüşümü
        if t == "MO" and family != "B737MAX":
            yr_val = convert_mo_to_yr(v)
            if yr_val is None:
                continue

            results.append({
                "type": "YR",
                "value": yr_val,
                "source": source_name,
                "raw_type": "MO",
                "raw_value": v,
                "exceed": interval_exceeds("YR", yr_val, aircraft_family)
            })

        else:
            results.append({
                "type": t,
                "value": v,
                "source": source_name,
                "raw_type": t,
                "raw_value": v,
                "exceed": interval_exceeds(t, v, aircraft_family)
            })

    return results

def deduplicate_intervals(intervals):
    seen = set()
    unique = []

    for x in intervals:
        key = (
            x.get("type"),
            round(float(x.get("value", 0)), 6),
            x.get("source"),
            x.get("raw_type"),
            round(float(x.get("raw_value", 0)), 6),
        )
        if key not in seen:
            seen.add(key)
            unique.append(x)

    return unique

def collect_labeled_section(lines, label):
    chunks = []
    n = len(lines)

    for i, line in enumerate(lines):
        up = normalize_text_for_search(line)

        if label not in up:
            continue

        buf = [up]

        for j in range(i + 1, min(i + 8, n)):
            nxt = normalize_text_for_search(lines[j])
            if not nxt:
                continue

            if j != i + 1 and any(sw in nxt for sw in STOP_WORDS):
                break

            buf.append(nxt)

        chunks.append(" ".join(buf))

    return chunks

def extract_intervals_from_page(page_text, aircraft_family=None):
    intervals = []
    if not page_text:
        return intervals

    lines = str(page_text).splitlines()

    threshold_chunks = collect_labeled_section(lines, "THRESHOLD")
    repeat_chunks = collect_labeled_section(lines, "REPEAT")

    for chunk in threshold_chunks:
        intervals.extend(extract_intervals_from_chunk(chunk, "THRESHOLD", aircraft_family))

    for chunk in repeat_chunks:
        intervals.extend(extract_intervals_from_chunk(chunk, "REPEAT", aircraft_family))

    if not intervals:
        txt = normalize_text_for_search(page_text)
        intervals.extend(extract_intervals_from_chunk(txt, "PAGE_FALLBACK", aircraft_family))

    return deduplicate_intervals(intervals)

def is_task_card_like_page(page_text: str) -> bool:
    txt = normalize_text_for_search(page_text)
    keywords = [
        "BOEING CARD NO",
        "THRESHOLD",
        "REPEAT",
        "ACCESS",
        "ZONE",
        "WORK AREA"
    ]
    score = sum(1 for k in keywords if k in txt)
    return score >= 2

def build_card_interval_map(pdf_bytes, aircraft_family=None):
    card_map = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            txt = normalize_text_for_search(text)

            if not txt:
                continue

            card_numbers = re.findall(r"\b\d{2}-\d{3}-\d{2}-\d{2}\b", txt)
            if not card_numbers:
                card_numbers = re.findall(r"\b\d{2}-\d{3}-\d{2}\b", txt)

            if not card_numbers:
                continue

            if not is_task_card_like_page(txt):
                continue

            intervals = extract_intervals_from_page(text, aircraft_family)
            if not intervals:
                continue

            unique_cards = sorted(set(card_numbers))
            for card_no in unique_cards:
                if card_no not in card_map:
                    card_map[card_no] = []
                card_map[card_no].extend(intervals)

    for card_no in list(card_map.keys()):
        card_map[card_no] = deduplicate_intervals(card_map[card_no])

    return card_map

def format_interval_summary(intervals):
    if not intervals:
        return ""

    parts = []
    for x in intervals:
        raw_type = x.get("raw_type", x.get("type", ""))
        raw_value = x.get("raw_value", x.get("value", ""))
        conv_type = x.get("type", "")
        conv_value = x.get("value", "")
        source = x.get("source", "")
        exceed = "Y" if x.get("exceed") else "N"

        if raw_type == "MO":
            part = f"{source}:{format_num(raw_value)}MO=>{format_num(conv_value)}YR({exceed})"
        else:
            part = f"{source}:{format_num(conv_value)}{conv_type}({exceed})"

        parts.append(part)

    return " | ".join(parts)

def get_interval_rule_text(aircraft_family):
    family = (aircraft_family or "").upper().strip()

    if family == "B737MAX":
        return "Interval limitleri (B737MAX): FH > 1600 | DY > 120 | MO > 4 | YR eşik dışı"
    elif family == "B737NG":
        return "Interval limitleri (B737NG): FH ≥ 15000 | FC ≥ 4500 | YR ≥ 3 | MO -> YR (12 MO = 1 YR)"
    return "Interval limitleri: Uçak tipi tanınamadı, varsayılan olarak NG kuralları uygulanır."

# -----------------------------
# Engineering mapping
# -----------------------------
def load_engineering_mapping(uploaded_excel):
    df = pd.read_excel(uploaded_excel)
    cols = {str(c).strip().upper(): c for c in df.columns}

    desc_col_key = next((k for k in cols.keys() if k.strip().upper() == "DESCRIPTION"), None)
    if not desc_col_key:
        raise KeyError("Mühendislik değerlendirmesi Excel'inde DESCRIPTION sütunu yok.")

    desc_col = cols[desc_col_key]
    cmt_col = cols.get("CMT")
    imt_col = cols.get("IMT")
    cdccl_col = cols.get("CDCCL")
    kompleks_col = cols.get("KOMPLEKS")

    mapping = {}
    kompleks_any = False

    for _, r in df.iterrows():
        key = clean_text_key(r.get(desc_col))
        if not key:
            continue

        cmt = yn_from_any(r.get(cmt_col)) if cmt_col else "N"
        imt = yn_from_any(r.get(imt_col)) if imt_col else "N"
        cdccl = yn_from_any(r.get(cdccl_col)) if cdccl_col else "N"
        kompleks = yn_from_any(r.get(kompleks_col)) if kompleks_col else "N"

        if kompleks == "Y":
            kompleks_any = True

        prev = mapping.get(key, {"cmt": "N", "imt": "N", "cdccl": "N"})
        mapping[key] = {
            "cmt": "Y" if (prev["cmt"] == "Y" or cmt == "Y") else "N",
            "imt": "Y" if (prev["imt"] == "Y" or imt == "Y") else "N",
            "cdccl": "Y" if (prev["cdccl"] == "Y" or cdccl == "Y") else "N",
        }

    return mapping, kompleks_any

# -----------------------------
# Fill template
# -----------------------------
def fill_template_excel(template_bytes, aircraft, package_name, tasks, wo_number):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_map = {norm_header(c.value): c.column for c in ws[1] if c.value}

    def col(name: str):
        return header_map.get(norm_header(name))

    required = [
        "Aircraf", "Check", "wo", "chapter", "sectIon", "task_card_descrIptIon",
        "addItIon_work", "edItor_used", "source_code", "rII", "cdccl",
        "crItIcal_task", "etops", "mechanIc", "skIll", "man_hours", "men_requIred"
    ]
    missing = [k for k in required if col(k) is None]
    if missing:
        raise KeyError(f"Şablon Excel'de şu başlıklar eksik: {missing}")

    start_row = 2
    for i, t in enumerate(tasks):
        r = start_row + i

        ws.cell(r, col("Aircraf")).value = aircraft
        ws.cell(r, col("Check")).value = package_name
        ws.cell(r, col("wo")).value = wo_number
        ws.cell(r, col("chapter")).value = 5
        ws.cell(r, col("sectIon")).value = 0
        ws.cell(r, col("task_card_descrIptIon")).value = t["description"]
        ws.cell(r, col("addItIon_work")).value = "YES"
        ws.cell(r, col("edItor_used")).value = "STYLESHEET"
        ws.cell(r, col("source_code")).value = "R"

        ws.cell(r, col("rII")).value = t["rII"]
        ws.cell(r, col("cdccl")).value = t["cdccl"]
        ws.cell(r, col("crItIcal_task")).value = t["critical_task"]
        ws.cell(r, col("etops")).value = "N"

        ws.cell(r, col("mechanIc")).value = "Y"
        ws.cell(r, col("skIll")).value = t["skill"]
        ws.cell(r, col("man_hours")).value = t["man_hour"]
        ws.cell(r, col("men_requIred")).value = 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# -----------------------------
# Export TSV
# -----------------------------
def workbook_bytes_to_tsv_bytes(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    headers = [safe_cell_str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    lines = ["\t".join(headers)]

    for r in range(2, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            row_vals.append(safe_cell_str(ws.cell(r, c).value))
        lines.append("\t".join(row_vals))

    return ("\n".join(lines)).encode("utf-8")

# -----------------------------
# Main action
# -----------------------------
if st.button("Excel Oluştur"):
    if not (pdf_file and template_file and wo_number):
        st.error("PDF, Excel şablon ve W/O numarasını girmen gerekiyor.")
    elif use_engineering and map_file is None:
        st.error("Mühendislik değerlendirmesi seçildi ama Excel yüklenmedi.")
    else:
        try:
            pdf_bytes = pdf_file.getvalue()
            template_bytes = template_file.getvalue()

            family, msg = detect_aircraft_family_from_cover(pdf_bytes)
            if family == "B737MAX":
                st.info(msg)
            elif family == "B737NG":
                st.success(msg)
            else:
                st.warning(msg)

            st.info(get_interval_rule_text(family))

            aircraft, package_name, tasks = extract_summary_tasks(pdf_bytes)
            card_interval_map = build_card_interval_map(pdf_bytes, family)

            interval_found_count = 0
            interval_exceed_count = 0

            for t in tasks:
                card_no = t.get("card_no", "")
                intervals = card_interval_map.get(card_no, []) if card_no else []

                t["intervals"] = intervals
                if intervals:
                    t["interval_exceed"] = "Y" if all(x.get("exceed") for x in intervals) else "N"
                else:
                    t["interval_exceed"] = "N"
                t["interval_summary"] = format_interval_summary(intervals)

                if intervals:
                    interval_found_count += 1
                if t["interval_exceed"] == "Y":
                    interval_exceed_count += 1

            location = get_location_from_package(package_name)
            mpd_cmr_interval = mpd_cmr_interval(tasks, "52-360-00-01")
            NDT_control=NDT_control(tasks, "EOD-B737-53-0010")
            access_issue=access_issue(tasks, "EOD-B737-51-0010")
            access_issue2=access_issue2(tasks, "55-826-01-01")
            access_issue3=access_issue3(tasks, "55-840-02-01")
            if mpd_cmr_interval:
                st.warning("52-360-00-01|‼️Kartın İntervali limit dışı fakat Special notunda limit içi olabilir. Kontrol edilmeli.‼️")
            if NDT_control:
                st.warning("EOD-B737-53-0010|‼️NDT Kontrolü olması sebebiyle Furkan Erence Cancel talebi yapmıştı.  kontrol lütfen.‼️")
            if access_issue:
                st.warning("EOD-B737-51-0010|‼️Detaylı gövde erişimi sorunu sebebiyle Furkan Erence Cancel talebi yapmıştı. kontrol lütfen.‼️")
            if access_issue2:
                st.warning("55-826-01-01|‼️Bölge erişimi sorunu sebebiyle Furkan Erence Cancel talebi yapmıştı. kontrol lütfen.‼️")
            if access_issue3:
                st.warning("55-840-02-01|‼️Bölge erişimi sorunu sebebiyle Furkan Erence Cancel talebi yapmıştı. kontrol lütfen.‼️")
            # MAX + ADB + EOD-B737-73-0003 uyarısı
            has_max_eod_card = has_eod_max_engine_run_card(tasks, "EOD-B737-73-0003")
            if family == "B737MAX" and location == "ADB" and has_max_eod_card:
                st.warning("EOD-B737-73-0003|‼️İzmirde MAX motor çalıştırma yetkili personel yoktu, teyit lazım.‼️")
            if use_engineering and map_file is not None:
                mapping, kompleks_any = load_engineering_mapping(map_file)

                for t in tasks:
                    k = clean_text_key(t["match_key"])
                    if k in mapping:
                        t["rII"] = mapping[k]["cmt"]
                        t["critical_task"] = mapping[k]["imt"]
                        t["cdccl"] = mapping[k]["cdccl"]
                    else:
                        t["rII"] = "N"
                        t["critical_task"] = "N"
                        t["cdccl"] = "N"

                if kompleks_any:
                    st.warning("⚠️ Kompleks iş var (KOMPLEKS=Y/YES bulundu)")

            total_mh = 0
            for t in tasks:
                try:
                    total_mh += int(str(t["man_hour"]).strip())
                except Exception:
                    pass

            c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
            c1.metric("Toplam İş", len(tasks))
            c2.metric("Toplam Man Hour", total_mh)
            c3.metric("rII=Y", sum(1 for t in tasks if t["rII"] == "Y"))
            c4.metric("Critical=Y", sum(1 for t in tasks if t["critical_task"] == "Y"))
            c5.metric("Lokasyon", location or "-")
            c6.metric("Interval Bulunan", interval_found_count)
            c7.metric("Limit Aşan", interval_exceed_count)

            if len(tasks) == 0:
                st.error("Summary tablosundan hiç iş çekilemedi.")
            else:
                interval_rows = []
                for t in tasks:
                    if t.get("card_no"):
                        interval_rows.append({
                            "Card No": t.get("card_no", ""),
                            "Description": t.get("match_key", ""),
                            "Interval Summary": t.get("interval_summary", ""),
                            "Interval Exceed": t.get("interval_exceed", "N")
                        })

                if interval_rows:
                    st.subheader("Interval Analizi")
                    interval_df = pd.DataFrame(interval_rows)
                    st.dataframe(interval_df, use_container_width=True)

                    exceed_df = interval_df[interval_df["Interval Exceed"] == "Y"].copy()
                    if not exceed_df.empty:
                        st.subheader("Limit Aşan Kartlar")
                        st.dataframe(exceed_df, use_container_width=True)

                filled_xlsx = fill_template_excel(
                    template_bytes=template_bytes,
                    aircraft=aircraft,
                    package_name=package_name,
                    tasks=tasks,
                    wo_number=wo_number
                )
                tsv_bytes = workbook_bytes_to_tsv_bytes(filled_xlsx)

                st.session_state["filled_xlsx"] = filled_xlsx
                st.session_state["filled_tsv"] = tsv_bytes
                st.session_state["dl_aircraft"] = aircraft
                st.session_state["dl_v"] += 1

                st.success("Dosyalar hazır. Aşağıdan indirebilirsin ✅")

        except Exception as e:
            st.error(f"Hata: {e}")

# -----------------------------
# Persistent download buttons
# -----------------------------
if st.session_state["filled_xlsx"] is not None:
    aircraft = st.session_state["dl_aircraft"] or "IMPORT"
    v = st.session_state["dl_v"]

    colA, colB = st.columns(2)
    with colA:
        st.download_button(
            label="IMPORT Excel (.xlsx)",
            data=st.session_state["filled_xlsx"],
            file_name=f"{aircraft} IMPORT EXCELI_v{v}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_xlsx_persist_{v}",
        )
    with colB:
        st.download_button(
            label="IMPORT Text (Tab Delimited .txt)",
            data=st.session_state["filled_tsv"],
            file_name=f"{aircraft} IMPORT EXCELI_v{v}.txt",
            mime="text/plain",
            key=f"dl_txt_persist_{v}",
        )




